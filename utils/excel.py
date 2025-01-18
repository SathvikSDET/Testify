
import openpyxl 

class Excel:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.loadbook_workbook(file_path)
    
    def get_sheet(self, sheet_name):
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            raise ValueError(f"Sheet {sheet_name} not found.")

    def get_Cells(self, rows ,columns):
        self.sheet = self.workbook.get_sheet()
        self.cell = self.sheets.cell(row = rows, column = columns)
        return self.cell.value
    
    def write_to_cell(self, sheet_name, row, column, value):
        sheet = self.get_sheet(sheet_name)
        cell = sheet.cell(row=row, column=column)
        cell.value = value
        self.workbook.save(self.file_path)  
       